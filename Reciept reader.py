import google.generativeai as genai
import os
import warnings
import easyocr
from PIL import Image
import pandas as pd
import cv2
import csv
import io
import itertools
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# Configure Google Generative AI
genai.configure(api_key="API KEY")

# Extract text using OCR
def extract_text(image_path):
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image file not found: {image_path}")

    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"Failed to load image: {image_path}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=FutureWarning)
        reader = easyocr.Reader(['en'])
    
    result = reader.readtext(image)
    text = ' '.join([res[1] for res in result])
    return str(text)

# Process the generated dictionary
def process_dict(res):
    data = ""
    res = itertools.islice(res, 0, None)
    
    for char in res:
        if char != "}":
            if char == "{":
                data += char
                continue
            if "{" in data:
                data += char
        else:
            data += "}"
            break
            
    print("DATA: ", data)
    return json.loads(data)

# Append data to Excel
def append_data(path):
    text = extract_text(path)
    print("\nExtracted Text:\n", text)
    
    for _ in range(5):
        try:
            # Generate structured content from extracted text using generative AI
            model = genai.GenerativeModel("gemini-1.5-flash")
            result = model.generate_content(f'Extract and structure the following receipt text into a python dictionary format with columns, include no comments in code: Date(%d-%m-%Y), Item(string), Quantity(int), Price(int), and location(string):\n\n"{text}"\n\n')

            result_text = result.text
            print("\nGenerated Output:\n", result_text)

            result = process_dict(result_text)

            date_formats = ["%d-%m-%Y", "%m-%d-%Y", "%d/%m/%Y", "%d-%b-%Y", "%b-%d-%Y"]

            current_month_year = None
            for fmt in date_formats:
                try:
                    current_month_year = datetime.strptime(result["Date"], fmt).strftime("%Y-%m")
                    break
                except ValueError:
                    continue

            if current_month_year is None:
                raise ValueError("Failed to parse date from receipt")

            print(f"Parsed Month-Year: {current_month_year}")

            excel_file = "budget.xlsx"
            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
            else:
                wb = load_workbook(excel_file)

            new_row = [result[i] for i in result]

            try:
            except KeyError:
                headers = ["Date", "Item", "Quantity", "Price", "Location"]
                ws = wb.create_sheet(title=current_month_year)
                ws.append(headers)


                for col_num, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.border = thin_border
                    cell.font = Font(bold=True)

            ws.append(new_row)
            wb.save(excel_file)
            break
            
        except ValueError:
            continue

if __name__ == "__main__":
    receipt_dir = "reciepts"
    for filename in os.listdir(receipt_dir):
        file_path = os.path.join(receipt_dir, filename)
        if os.path.isfile(file_path):
            append_data(file_path)
